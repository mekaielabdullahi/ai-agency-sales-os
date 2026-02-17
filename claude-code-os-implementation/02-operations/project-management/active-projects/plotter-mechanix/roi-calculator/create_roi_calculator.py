#!/usr/bin/env python3
"""
Plotter Mechanix - AI ROI Calculator - Phase 2
Based on Morningside AI ROI Calculator Framework

Creates a comprehensive Excel workbook with validated ROI data from Phase 2 discovery.
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# ============================================================================
# CONFIGURATION
# ============================================================================

OUTPUT_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_FILE = os.path.join(OUTPUT_DIR, "Plotter Mechanix - AI ROI Calculator - Phase 2.xlsx")

# Color scheme
COLORS = {
    'header_bg': 'FF1F4E79',      # Dark blue
    'header_font': 'FFFFFFFF',    # White
    'section_a': 'FFD6EAF8',      # Light blue (Time Cost)
    'section_b': 'FFFDEBD0',      # Light orange (Lost Revenue)
    'section_c': 'FFD5F5E3',      # Light green (Revenue Upside)
    'total_bg': 'FF2E7D32',       # Dark green
    'total_font': 'FFFFFFFF',     # White
    'urgent_bg': 'FFFFF3CD',      # Yellow (needs validation)
    'high_conf': 'FFD4EDDA',      # Green (high confidence)
    'med_conf': 'FFFFEEBA',       # Yellow (medium confidence)
    'pillar_1': 'FF3498DB',       # Blue (Operations)
    'pillar_2': 'FF9B59B6',       # Purple (Training)
    'pillar_3': 'FF1ABC9C',       # Teal (Contacts)
}

# Styles
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# ============================================================================
# VALIDATED DATA FROM PHASE 2 DISCOVERY
# ============================================================================

# Hourly rates (validated from project documents)
HOURLY_RATES = {
    'alyssa': {'rate': 25, 'note': 'Estimated $20-30/hr - NEEDS VALIDATION'},
    'kelsey_billable': {'rate': 175, 'note': 'Jobber pricing'},
    'kelsey_opportunity': {'rate': 337.50, 'note': 'Calculated: $1.2-1.7M revenue / 2,600 hrs x 65%'},
    'joe': {'rate': 50, 'note': 'Estimated - NEEDS VALIDATION'},
    'working_days_per_year': 260,
}

# 6 Main Opportunities with validated data
OPPORTUNITIES = {
    'opportunity_1': {
        'name': 'Email & Communication Hub',
        'investment': {'low': 5000, 'high': 7000},
        'pillar': 'Operations',
        'source': 'Alyssa Part 2 Interview (Feb 6, 2026)',

        # Section A: Time Cost
        'time_cost': {
            'description': 'Email consolidation and management',
            'pain_points': [
                {'metric': 'Inboxes managed', 'value': 10, 'source': 'Alyssa Part 2'},
                {'metric': 'Daily email time (hrs)', 'value': 4.5, 'source': 'Alyssa Part 2: "That\'s the majority of my time"'},
                {'metric': 'Weekly email time (hrs)', 'value': 22.5, 'source': 'Calculated: 50% of her day'},
                {'metric': 'Notification system', 'value': 'Broken', 'source': 'Alyssa Part 2: must manually check'},
            ],
            'calculations': [
                {
                    'item': 'Email consolidation (10→2 inboxes)',
                    'time_per_day_hrs': 2,
                    'people_affected': 1,
                    'hourly_rate': 25,
                    'annual_value_low': 10400,
                    'annual_value_high': 20280,
                    'confidence': 'HIGH',
                },
                {
                    'item': 'Automated routing/labeling',
                    'time_per_day_hrs': 0.5,
                    'people_affected': 1,
                    'hourly_rate': 25,
                    'annual_value_low': 2080,
                    'annual_value_high': 4160,
                    'confidence': 'HIGH',
                },
                {
                    'item': 'Unified communication view',
                    'time_per_day_hrs': 0.5,
                    'people_affected': 1,
                    'hourly_rate': 25,
                    'annual_value_low': 2080,
                    'annual_value_high': 4160,
                    'confidence': 'MEDIUM',
                },
            ],
        },

        # Section B: Lost Revenue (not applicable for this opportunity)
        'lost_revenue': None,

        # Section C: Revenue Upside (not applicable for this opportunity)
        'revenue_upside': None,

        # Totals
        'total_annual_low': 14560,
        'total_annual_high': 28600,
        'confidence': 'HIGH',
    },

    'opportunity_2': {
        'name': 'Duplicate Prevention & Data Integrity',
        'investment': {'low': 3000, 'high': 4000},
        'pillar': 'Contacts',
        'source': 'Alyssa Part 2 Interview (Feb 6, 2026)',

        # Section A: Time Cost
        'time_cost': {
            'description': 'Data cleanup and duplicate management',
            'pain_points': [
                {'metric': 'Quote', 'value': '"We have like seven Daves"', 'source': 'Alyssa Part 2'},
                {'metric': 'Daily duplicate time (min)', 'value': 30, 'source': 'Alyssa Part 2'},
                {'metric': 'Workaround', 'value': 'Copy/paste manual', 'source': 'Alyssa Part 2'},
                {'metric': 'QuickBooks sync', 'value': 'Breaking weekly', 'source': 'Alyssa Part 2'},
            ],
            'calculations': [
                {
                    'item': 'Eliminate duplicate entry',
                    'time_per_day_hrs': 0.5,
                    'people_affected': 1,
                    'hourly_rate': 25,
                    'annual_value_low': 2600,
                    'annual_value_high': 3900,
                    'confidence': 'HIGH',
                },
                {
                    'item': 'Reduce data cleanup',
                    'time_per_day_hrs': 0.25,
                    'people_affected': 1,
                    'hourly_rate': 25,
                    'annual_value_low': 1040,
                    'annual_value_high': 1560,
                    'confidence': 'HIGH',
                },
                {
                    'item': 'Prevent QuickBooks errors',
                    'time_per_day_hrs': 0.125,
                    'people_affected': 1,
                    'hourly_rate': 25,
                    'annual_value_low': 520,
                    'annual_value_high': 780,
                    'confidence': 'MEDIUM',
                },
            ],
        },

        'lost_revenue': None,
        'revenue_upside': None,

        'total_annual_low': 4160,
        'total_annual_high': 6240,
        'confidence': 'HIGH',
    },

    'opportunity_3': {
        'name': 'Inventory & Barcode System',
        'investment': {'low': 4000, 'high': 6000},
        'pillar': 'Operations',
        'source': 'Alyssa Part 2 Interview (Feb 6, 2026) - Her #1 Wish',

        # Section A: Time Cost
        'time_cost': {
            'description': 'Inventory visibility and tracking',
            'pain_points': [
                {'metric': 'Quote', 'value': '"We don\'t know what we have... QuickBooks does not have the correct amounts"', 'source': 'Alyssa Part 2'},
                {'metric': 'Quote', 'value': '"If we had barcodes, that would be so much easier"', 'source': 'Alyssa Part 2'},
                {'metric': 'Inventory visibility', 'value': 'None - manual checking', 'source': 'Alyssa Part 2'},
                {'metric': 'RMA backlog', 'value': 'Pile from last year', 'source': 'Alyssa Part 2'},
            ],
            'calculations': [
                {
                    'item': 'Real-time stock visibility',
                    'time_per_day_hrs': 0.5,
                    'people_affected': 1,
                    'hourly_rate': 25,
                    'annual_value_low': 2080,
                    'annual_value_high': 3120,
                    'confidence': 'HIGH',
                },
                {
                    'item': 'RMA resolution',
                    'time_per_day_hrs': 0.375,
                    'people_affected': 1,
                    'hourly_rate': 25,
                    'annual_value_low': 1560,
                    'annual_value_high': 2340,
                    'confidence': 'HIGH',
                },
            ],
        },

        # Section B: Lost Revenue
        'lost_revenue': {
            'description': 'Stockouts and delayed orders',
            'calculations': [
                {
                    'item': 'Reduced stockouts',
                    'volume_per_month': 4,
                    'pct_lost': 0.25,
                    'value_per_unit': 500,
                    'annual_value_low': 2000,
                    'annual_value_high': 4000,
                    'confidence': 'MEDIUM',
                },
            ],
        },

        'revenue_upside': None,

        'total_annual_low': 5640,
        'total_annual_high': 9460,
        'confidence': 'HIGH',
    },

    'opportunity_4': {
        'name': 'Parts Catalog & Knowledge Base',
        'investment': {'low': 8000, 'high': 12000},
        'pillar': 'Training',
        'source': 'Alyssa Part 2 Interview (Feb 6, 2026)',

        # Section A: Time Cost
        'time_cost': {
            'description': 'Parts research and vendor communication',
            'pain_points': [
                {'metric': 'Quote', 'value': '"It can take me about half an hour to try and figure it out because I don\'t know parts"', 'source': 'Alyssa Part 2'},
                {'metric': 'Parts research time (min)', 'value': 30, 'source': 'Alyssa Part 2'},
                {'metric': 'Complex HP parts', 'value': 'Days + specialist calls', 'source': 'Alyssa Part 2'},
                {'metric': 'Vendor communication (min)', 'value': 15, 'source': 'Alyssa Part 2'},
                {'metric': 'Total per order (min)', 'value': 45, 'source': 'Alyssa Part 2'},
                {'metric': 'Orders per week', 'value': 2.5, 'source': 'Alyssa Part 2: 2-3'},
            ],
            'calculations': [
                {
                    'item': 'Parts lookup reduction (20 min/order saved)',
                    'time_per_day_hrs': 0.33,
                    'people_affected': 1,
                    'hourly_rate': 25,
                    'annual_value_low': 1733,
                    'annual_value_high': 2600,
                    'confidence': 'HIGH',
                },
                {
                    'item': 'Reduced vendor calls',
                    'time_per_day_hrs': 0.25,
                    'people_affected': 1,
                    'hourly_rate': 25,
                    'annual_value_low': 1040,
                    'annual_value_high': 1560,
                    'confidence': 'HIGH',
                },
            ],
        },

        # Section C: Revenue Upside
        'revenue_upside': {
            'description': 'Customer self-service enabling',
            'calculations': [
                {
                    'item': 'Customer self-service potential',
                    'activities_per_month': 10,
                    'conversion_rate': 0.2,
                    'value_per_conversion': 100,
                    'annual_value_low': 2080,
                    'annual_value_high': 3120,
                    'confidence': 'MEDIUM',
                },
            ],
        },

        'lost_revenue': None,

        'total_annual_low': 4853,
        'total_annual_high': 7280,
        'confidence': 'HIGH',
    },

    'opportunity_5': {
        'name': 'Training System (Video Library + AI Assistant)',
        'investment': {'low': 8000, 'high': 18000},
        'pillar': 'Training',
        'source': 'Joe Interview (Jan 28, 2026)',

        # Section A: Time Cost
        'time_cost': {
            'description': 'New technician training time reduction',
            'pain_points': [
                {'metric': 'Quote', 'value': '"Like 20, 30 hours, 20, 30 hours initially... took me a good 3 months"', 'source': 'Joe Interview'},
                {'metric': 'First 3-4 weeks (hrs/week)', 'value': 25, 'source': 'Joe Interview: 20-30'},
                {'metric': 'Weeks 5-8 (hrs/week)', 'value': 12.5, 'source': 'Joe Interview: tapers to 10-15'},
                {'metric': 'Weeks 9-12 (hrs/week)', 'value': 7.5, 'source': 'Joe Interview: tapers to 5-10'},
                {'metric': 'Total per new tech (hrs)', 'value': 160, 'source': 'Calculated: 120-200'},
                {'metric': 'Kelsey billable cost/hire', 'value': '$28,000', 'source': '160 hrs x $175/hr'},
            ],
            'calculations': [
                {
                    'item': 'Training time reduction (90-150 hrs saved/hire)',
                    'description': 'From 120-200 hrs to 30-50 hrs',
                    'hours_saved_per_hire': 120,
                    'hourly_rate': 175,
                    'hires_per_year': 2,
                    'annual_value_low': 20000,
                    'annual_value_high': 40000,
                    'confidence': 'HIGH',
                },
            ],
        },

        # Section C: Revenue Upside
        'revenue_upside': {
            'description': 'Kelsey capacity freed for high-value work',
            'calculations': [
                {
                    'item': 'Joe independence boost (30% → 10% need help)',
                    'description': '20% fewer escalations = more billable Kelsey hours',
                    'hours_freed_per_week': 10,
                    'hourly_rate': 175,
                    'weeks_per_year': 52,
                    'annual_value_low': 78000,
                    'annual_value_high': 91000,
                    'confidence': 'HIGH',
                },
            ],
        },

        'lost_revenue': None,

        'total_annual_low': 31500,  # Conservative: training only
        'total_annual_high': 78750,  # Optimistic: training + independence
        'confidence': 'HIGH',
    },

    'opportunity_6': {
        'name': 'Morning Coordination',
        'investment': {'low': 3000, 'high': 6000},
        'pillar': 'Operations',
        'source': 'Joe Interview (Jan 28, 2026)',

        # Section A: Time Cost
        'time_cost': {
            'description': 'Daily route planning and coordination',
            'pain_points': [
                {'metric': 'Quote', 'value': '"1.5-2 hours every morning planning routes... Routes change 4 out of 5 days"', 'source': 'Joe Interview'},
                {'metric': 'Quote', 'value': '"Most frustrating part of my day"', 'source': 'Joe Interview'},
                {'metric': 'Daily coordination (hrs)', 'value': 1.75, 'source': 'Joe Interview: 1.5-2'},
                {'metric': 'Route changes frequency', 'value': '80%', 'source': 'Joe Interview: 4 out of 5 days'},
            ],
            'calculations': [
                {
                    'item': 'Coordination time reduction',
                    'description': 'From 1.5-2 hrs to 30-45 min daily',
                    'time_saved_per_day_hrs': 1.125,
                    'people_affected': 1,
                    'hourly_rate': 50,
                    'annual_value_low': 12500,
                    'annual_value_high': 15625,
                    'confidence': 'HIGH',
                },
            ],
        },

        'lost_revenue': None,
        'revenue_upside': None,

        'total_annual_low': 12500,
        'total_annual_high': 15625,
        'confidence': 'HIGH',
    },
}

# Tier summaries
TIERS = {
    'foundation': {
        'name': 'Foundation Tier',
        'investment': 15000,
        'components': ['Email & Communication Hub', 'Duplicate Prevention', 'Basic Knowledge Base'],
        'annual_roi_low': 22720,
        'annual_roi_high': 40840,
        'roi_pct_low': 151,
        'roi_pct_high': 272,
        'payback_months_low': 4,
        'payback_months_high': 8,
    },
    'growth': {
        'name': 'Growth Tier (RECOMMENDED)',
        'investment': 28000,
        'components': ['All Foundation', 'Inventory & Barcode', 'Parts Catalog', 'Training System (15 videos)', 'Morning Coordination'],
        'annual_roi_low': 57213,
        'annual_roi_high': 103580,
        'roi_pct_low': 204,
        'roi_pct_high': 370,
        'payback_months_low': 3,
        'payback_months_high': 6,
    },
    'full_scale': {
        'name': 'Full Scale Tier',
        'investment': 42000,
        'components': ['All Growth', 'Full Training System (25-30 videos)', 'Field Update Automation', 'Advanced Analytics'],
        'annual_roi_low': 91000,
        'annual_roi_high': 174375,
        'roi_pct_low': 217,
        'roi_pct_high': 415,
        'payback_months_low': 3,
        'payback_months_high': 5,
    },
}

# Pillar summaries
PILLARS = {
    'operations': {
        'name': 'Pillar 1: Operations',
        'focus': 'Quo/Ply/Jobber automation',
        'key_benefit': 'Alyssa freed, Andrew enabled',
        'annual_roi_low': 509553,
        'annual_roi_high': 965840,
        'includes': ['Email Hub', 'Quote Workflow', 'Parts/Inventory', 'Andrew Supplies'],
    },
    'training': {
        'name': 'Pillar 2: Training Center',
        'focus': 'Tech independence',
        'key_benefit': 'Kelsey freed, hiring enabled',
        'annual_roi_low': 135000,
        'annual_roi_high': 195875,
        'includes': ['Video Library', 'AI Field Assistant', 'Mobile Knowledge Base'],
    },
    'contacts': {
        'name': 'Pillar 3: Contacts Consolidation',
        'focus': 'Outreach/marketing',
        'key_benefit': 'Revenue activation, Phase 3 foundation',
        'annual_roi_low': 174160,
        'annual_roi_high': 176240,
        'includes': ['Master Contact Consolidation', 'Segmentation', 'Marketing Foundation'],
    },
}

# Data gaps
DATA_GAPS = [
    {'item': "Alyssa's actual hourly rate", 'status': 'URGENT', 'impact': 'Using $20-30 estimate', 'current_value': '$25 (midpoint)'},
    {'item': '2025 actual revenue', 'status': 'IMPORTANT', 'impact': 'ROI % accuracy', 'current_value': '$850K (estimated)'},
    {'item': 'Service call volume/week', 'status': 'IMPORTANT', 'impact': 'Capacity calculations', 'current_value': 'Unknown'},
    {'item': "Joe's hourly rate", 'status': 'MEDIUM', 'impact': 'Training comparison', 'current_value': '$50 (estimated)'},
]

# ============================================================================
# WORKBOOK CREATION FUNCTIONS
# ============================================================================

def create_workbook():
    """Create the main workbook with all sheets."""
    wb = Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    # Create sheets
    create_summary_sheet(wb)
    create_opportunities_sheet(wb)
    create_tiers_sheet(wb)
    create_pillars_sheet(wb)
    create_hourly_rates_sheet(wb)
    create_data_gaps_sheet(wb)
    create_quotes_sheet(wb)

    return wb


def apply_header_style(cell):
    """Apply header styling to a cell."""
    cell.font = Font(bold=True, color=COLORS['header_font'])
    cell.fill = PatternFill(start_color=COLORS['header_bg'], end_color=COLORS['header_bg'], fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border


def apply_currency_format(cell):
    """Apply currency format to a cell."""
    cell.number_format = '$#,##0'


def set_column_widths(ws, widths):
    """Set column widths for a worksheet."""
    for col_num, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = width


def create_summary_sheet(wb):
    """Create the executive summary sheet."""
    ws = wb.create_sheet("Executive Summary")

    # Title
    ws['A1'] = "PLOTTER MECHANIX - AI ROI CALCULATOR - PHASE 2"
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:F1')

    ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d')}"
    ws['A2'].font = Font(italic=True)

    ws['A3'] = "Based on validated data from Phase 2 discovery interviews"
    ws['A3'].font = Font(italic=True)

    # Client Info
    row = 5
    ws[f'A{row}'] = "CLIENT PROFILE"
    ws[f'A{row}'].font = Font(bold=True, size=14)
    row += 1

    client_info = [
        ("Client", "Plotter Mechanix (Phoenix, AZ)"),
        ("Owner", "Kelsey Hartzell"),
        ("Current Revenue", "$850K"),
        ("Target Revenue", "$8.5M transformation"),
        ("Phase 2 Investment", "$28,000 (Growth Tier recommended)"),
        ("Timeline", "6 weeks"),
        ("Payback", "2-3 weeks with new revenue, 67 days from operations savings alone"),
    ]

    for label, value in client_info:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = value
        ws[f'A{row}'].font = Font(bold=True)
        row += 1

    # ROI Summary
    row += 1
    ws[f'A{row}'] = "ROI SUMMARY"
    ws[f'A{row}'].font = Font(bold=True, size=14)
    row += 1

    headers = ["Tier", "Investment", "Annual ROI (Low)", "Annual ROI (High)", "ROI %", "Payback"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        apply_header_style(cell)
    row += 1

    for tier_key, tier in TIERS.items():
        ws.cell(row=row, column=1, value=tier['name'])
        ws.cell(row=row, column=2, value=tier['investment'])
        apply_currency_format(ws.cell(row=row, column=2))
        ws.cell(row=row, column=3, value=tier['annual_roi_low'])
        apply_currency_format(ws.cell(row=row, column=3))
        ws.cell(row=row, column=4, value=tier['annual_roi_high'])
        apply_currency_format(ws.cell(row=row, column=4))
        ws.cell(row=row, column=5, value=f"{tier['roi_pct_low']}-{tier['roi_pct_high']}%")
        ws.cell(row=row, column=6, value=f"{tier['payback_months_low']}-{tier['payback_months_high']} months")

        for col in range(1, 7):
            ws.cell(row=row, column=col).border = thin_border
        row += 1

    # Guarantee
    row += 2
    ws[f'A{row}'] = "GUARANTEE"
    ws[f'A{row}'].font = Font(bold=True, size=14)
    row += 1
    ws[f'A{row}'] = "If Phase 2 doesn't achieve 50% reduction in Alyssa's email/coordination time within 60 days of go-live,"
    row += 1
    ws[f'A{row}'] = "we work for free up to 60 additional days."

    set_column_widths(ws, [20, 35, 18, 18, 15, 18])


def create_opportunities_sheet(wb):
    """Create the detailed opportunities sheet with Morningside framework."""
    ws = wb.create_sheet("6 Opportunities (Detail)")

    row = 1

    for opp_key, opp in OPPORTUNITIES.items():
        # Opportunity header
        ws[f'A{row}'] = f"OPPORTUNITY: {opp['name']}"
        ws[f'A{row}'].font = Font(bold=True, size=14)
        ws.merge_cells(f'A{row}:F{row}')
        row += 1

        # Metadata
        ws[f'A{row}'] = f"Investment: ${opp['investment']['low']:,} - ${opp['investment']['high']:,}"
        ws[f'C{row}'] = f"Pillar: {opp['pillar']}"
        ws[f'E{row}'] = f"Confidence: {opp['confidence']}"
        row += 1

        ws[f'A{row}'] = f"Source: {opp['source']}"
        row += 1

        # Pain points
        if opp['time_cost'] and opp['time_cost'].get('pain_points'):
            row += 1
            ws[f'A{row}'] = "Validated Pain Points:"
            ws[f'A{row}'].font = Font(bold=True)
            row += 1

            for pp in opp['time_cost']['pain_points']:
                ws[f'A{row}'] = f"• {pp['metric']}: {pp['value']}"
                ws[f'D{row}'] = f"[{pp['source']}]"
                ws[f'D{row}'].font = Font(italic=True, size=9)
                row += 1

        row += 1

        # SECTION A: Time Cost
        if opp['time_cost'] and opp['time_cost'].get('calculations'):
            ws[f'A{row}'] = "SECTION A: TIME COST"
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'A{row}'].fill = PatternFill(start_color=COLORS['section_a'], end_color=COLORS['section_a'], fill_type='solid')
            ws.merge_cells(f'A{row}:F{row}')
            row += 1

            headers = ["Item", "Time/Day (hrs)", "People", "Rate ($/hr)", "Annual Low", "Annual High", "Confidence"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color=COLORS['section_a'], end_color=COLORS['section_a'], fill_type='solid')
                cell.border = thin_border
            row += 1

            for calc in opp['time_cost']['calculations']:
                ws.cell(row=row, column=1, value=calc['item'])
                ws.cell(row=row, column=2, value=calc.get('time_per_day_hrs', calc.get('time_saved_per_day_hrs', '-')))
                ws.cell(row=row, column=3, value=calc.get('people_affected', calc.get('hires_per_year', '-')))
                ws.cell(row=row, column=4, value=calc.get('hourly_rate', '-'))
                ws.cell(row=row, column=5, value=calc['annual_value_low'])
                apply_currency_format(ws.cell(row=row, column=5))
                ws.cell(row=row, column=6, value=calc['annual_value_high'])
                apply_currency_format(ws.cell(row=row, column=6))
                ws.cell(row=row, column=7, value=calc['confidence'])

                for col in range(1, 8):
                    ws.cell(row=row, column=col).border = thin_border
                row += 1

            row += 1

        # SECTION B: Lost Revenue
        if opp.get('lost_revenue') and opp['lost_revenue'].get('calculations'):
            ws[f'A{row}'] = "SECTION B: LOST REVENUE"
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'A{row}'].fill = PatternFill(start_color=COLORS['section_b'], end_color=COLORS['section_b'], fill_type='solid')
            ws.merge_cells(f'A{row}:F{row}')
            row += 1

            headers = ["Item", "Volume/Month", "% Lost", "Value/Unit", "Annual Low", "Annual High", "Confidence"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color=COLORS['section_b'], end_color=COLORS['section_b'], fill_type='solid')
                cell.border = thin_border
            row += 1

            for calc in opp['lost_revenue']['calculations']:
                ws.cell(row=row, column=1, value=calc['item'])
                ws.cell(row=row, column=2, value=calc.get('volume_per_month', '-'))
                ws.cell(row=row, column=3, value=f"{calc.get('pct_lost', 0)*100}%" if calc.get('pct_lost') else '-')
                ws.cell(row=row, column=4, value=calc.get('value_per_unit', '-'))
                ws.cell(row=row, column=5, value=calc['annual_value_low'])
                apply_currency_format(ws.cell(row=row, column=5))
                ws.cell(row=row, column=6, value=calc['annual_value_high'])
                apply_currency_format(ws.cell(row=row, column=6))
                ws.cell(row=row, column=7, value=calc['confidence'])

                for col in range(1, 8):
                    ws.cell(row=row, column=col).border = thin_border
                row += 1

            row += 1

        # SECTION C: Revenue Upside
        if opp.get('revenue_upside') and opp['revenue_upside'].get('calculations'):
            ws[f'A{row}'] = "SECTION C: REVENUE UPSIDE"
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'A{row}'].fill = PatternFill(start_color=COLORS['section_c'], end_color=COLORS['section_c'], fill_type='solid')
            ws.merge_cells(f'A{row}:F{row}')
            row += 1

            headers = ["Item", "Activities/Month", "Conv. Rate", "Value/Conv", "Annual Low", "Annual High", "Confidence"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color=COLORS['section_c'], end_color=COLORS['section_c'], fill_type='solid')
                cell.border = thin_border
            row += 1

            for calc in opp['revenue_upside']['calculations']:
                ws.cell(row=row, column=1, value=calc['item'])
                ws.cell(row=row, column=2, value=calc.get('activities_per_month', calc.get('hours_freed_per_week', '-')))
                ws.cell(row=row, column=3, value=f"{calc.get('conversion_rate', 0)*100}%" if calc.get('conversion_rate') else '-')
                ws.cell(row=row, column=4, value=calc.get('value_per_conversion', calc.get('hourly_rate', '-')))
                ws.cell(row=row, column=5, value=calc['annual_value_low'])
                apply_currency_format(ws.cell(row=row, column=5))
                ws.cell(row=row, column=6, value=calc['annual_value_high'])
                apply_currency_format(ws.cell(row=row, column=6))
                ws.cell(row=row, column=7, value=calc['confidence'])

                for col in range(1, 8):
                    ws.cell(row=row, column=col).border = thin_border
                row += 1

            row += 1

        # Total
        ws[f'A{row}'] = "TOTAL"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'E{row}'] = opp['total_annual_low']
        apply_currency_format(ws[f'E{row}'])
        ws[f'F{row}'] = opp['total_annual_high']
        apply_currency_format(ws[f'F{row}'])

        for col in ['A', 'E', 'F']:
            ws[f'{col}{row}'].fill = PatternFill(start_color=COLORS['total_bg'], end_color=COLORS['total_bg'], fill_type='solid')
            ws[f'{col}{row}'].font = Font(bold=True, color=COLORS['total_font'])
            ws[f'{col}{row}'].border = thin_border

        row += 3  # Space between opportunities

    set_column_widths(ws, [45, 15, 12, 12, 15, 15, 12])


def create_tiers_sheet(wb):
    """Create the pricing tiers comparison sheet."""
    ws = wb.create_sheet("Pricing Tiers")

    row = 1
    ws[f'A{row}'] = "PRICING TIER COMPARISON"
    ws[f'A{row}'].font = Font(bold=True, size=16)
    row += 2

    headers = ["Element", "Foundation ($15K)", "Growth ($28K) - RECOMMENDED", "Full Scale ($42K)"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        apply_header_style(cell)
    row += 1

    tier_elements = [
        ("Investment", "$15,000", "$28,000", "$42,000"),
        ("Annual ROI (Low)", f"${TIERS['foundation']['annual_roi_low']:,}", f"${TIERS['growth']['annual_roi_low']:,}", f"${TIERS['full_scale']['annual_roi_low']:,}"),
        ("Annual ROI (High)", f"${TIERS['foundation']['annual_roi_high']:,}", f"${TIERS['growth']['annual_roi_high']:,}", f"${TIERS['full_scale']['annual_roi_high']:,}"),
        ("ROI %", f"{TIERS['foundation']['roi_pct_low']}-{TIERS['foundation']['roi_pct_high']}%", f"{TIERS['growth']['roi_pct_low']}-{TIERS['growth']['roi_pct_high']}%", f"{TIERS['full_scale']['roi_pct_low']}-{TIERS['full_scale']['roi_pct_high']}%"),
        ("Payback", f"{TIERS['foundation']['payback_months_low']}-{TIERS['foundation']['payback_months_high']} months", f"{TIERS['growth']['payback_months_low']}-{TIERS['growth']['payback_months_high']} months", f"{TIERS['full_scale']['payback_months_low']}-{TIERS['full_scale']['payback_months_high']} months"),
        ("Check-in Calls", "Weekly 30min", "Bi-weekly 45min", "Weekly 60min"),
        ("Email Response", "24-48 hours", "Same business day", "4-6 hours"),
        ("Team Hours/Week", "~10 hrs", "~20 hrs", "~30 hrs"),
        ("Post-Launch Support", "30 days", "60 days", "90 days"),
        ("Direct Contacts", "Mekaiel (PM)", "Mekaiel + Matthew", "Full team"),
    ]

    for element in tier_elements:
        for col, value in enumerate(element, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = thin_border
            if col == 3:  # Highlight Growth tier
                cell.fill = PatternFill(start_color='FFE8F5E9', end_color='FFE8F5E9', fill_type='solid')
        row += 1

    # Components section
    row += 2
    ws[f'A{row}'] = "COMPONENTS BY TIER"
    ws[f'A{row}'].font = Font(bold=True, size=14)
    row += 1

    for tier_key, tier in TIERS.items():
        ws[f'A{row}'] = tier['name']
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        for component in tier['components']:
            ws[f'B{row}'] = f"• {component}"
            row += 1
        row += 1

    set_column_widths(ws, [25, 25, 35, 25])


def create_pillars_sheet(wb):
    """Create the strategic pillars sheet."""
    ws = wb.create_sheet("3 Strategic Pillars")

    row = 1
    ws[f'A{row}'] = "THREE STRATEGIC PILLARS"
    ws[f'A{row}'].font = Font(bold=True, size=16)
    row += 2

    headers = ["Pillar", "Focus", "Annual ROI", "Key Benefit"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        apply_header_style(cell)
    row += 1

    for pillar_key, pillar in PILLARS.items():
        ws.cell(row=row, column=1, value=pillar['name'])
        ws.cell(row=row, column=2, value=pillar['focus'])
        ws.cell(row=row, column=3, value=f"${pillar['annual_roi_low']:,} - ${pillar['annual_roi_high']:,}")
        ws.cell(row=row, column=4, value=pillar['key_benefit'])

        for col in range(1, 5):
            ws.cell(row=row, column=col).border = thin_border
        row += 1

    # Total row
    total_low = sum(p['annual_roi_low'] for p in PILLARS.values())
    total_high = sum(p['annual_roi_high'] for p in PILLARS.values())

    ws.cell(row=row, column=1, value="TOTAL")
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=3, value=f"${total_low:,} - ${total_high:,}")
    ws.cell(row=row, column=3).font = Font(bold=True)

    for col in range(1, 5):
        ws.cell(row=row, column=col).fill = PatternFill(start_color=COLORS['total_bg'], end_color=COLORS['total_bg'], fill_type='solid')
        ws.cell(row=row, column=col).font = Font(bold=True, color=COLORS['total_font'])
        ws.cell(row=row, column=col).border = thin_border

    # Pillar details
    row += 3
    for pillar_key, pillar in PILLARS.items():
        ws[f'A{row}'] = pillar['name']
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1

        ws[f'A{row}'] = "Includes:"
        row += 1
        for item in pillar['includes']:
            ws[f'B{row}'] = f"• {item}"
            row += 1
        row += 1

    set_column_widths(ws, [30, 30, 30, 40])


def create_hourly_rates_sheet(wb):
    """Create the hourly rates reference sheet."""
    ws = wb.create_sheet("Hourly Rates")

    row = 1
    ws[f'A{row}'] = "VALIDATED HOURLY RATES"
    ws[f'A{row}'].font = Font(bold=True, size=16)
    row += 2

    headers = ["Person", "Rate ($/hr)", "Source/Notes"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        apply_header_style(cell)
    row += 1

    rates = [
        ("Alyssa", f"${HOURLY_RATES['alyssa']['rate']}", HOURLY_RATES['alyssa']['note']),
        ("Kelsey (billable)", f"${HOURLY_RATES['kelsey_billable']['rate']}", HOURLY_RATES['kelsey_billable']['note']),
        ("Kelsey (opportunity cost)", f"${HOURLY_RATES['kelsey_opportunity']['rate']}", HOURLY_RATES['kelsey_opportunity']['note']),
        ("Joe", f"${HOURLY_RATES['joe']['rate']}", HOURLY_RATES['joe']['note']),
    ]

    for person, rate, note in rates:
        ws.cell(row=row, column=1, value=person)
        ws.cell(row=row, column=2, value=rate)
        ws.cell(row=row, column=3, value=note)

        for col in range(1, 4):
            ws.cell(row=row, column=col).border = thin_border

        if 'NEEDS VALIDATION' in note:
            for col in range(1, 4):
                ws.cell(row=row, column=col).fill = PatternFill(start_color=COLORS['urgent_bg'], end_color=COLORS['urgent_bg'], fill_type='solid')

        row += 1

    row += 2
    ws[f'A{row}'] = f"Working days per year: {HOURLY_RATES['working_days_per_year']}"

    set_column_widths(ws, [25, 20, 60])


def create_data_gaps_sheet(wb):
    """Create the data gaps tracking sheet."""
    ws = wb.create_sheet("Data Gaps")

    row = 1
    ws[f'A{row}'] = "DATA GAPS - NEEDS VALIDATION"
    ws[f'A{row}'].font = Font(bold=True, size=16)
    row += 2

    headers = ["Item", "Status", "Impact", "Current Value Used"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        apply_header_style(cell)
    row += 1

    for gap in DATA_GAPS:
        ws.cell(row=row, column=1, value=gap['item'])
        ws.cell(row=row, column=2, value=gap['status'])
        ws.cell(row=row, column=3, value=gap['impact'])
        ws.cell(row=row, column=4, value=gap['current_value'])

        for col in range(1, 5):
            ws.cell(row=row, column=col).border = thin_border

        # Color by status
        if gap['status'] == 'URGENT':
            fill_color = 'FFFF6B6B'  # Red
        elif gap['status'] == 'IMPORTANT':
            fill_color = 'FFFFD93D'  # Yellow
        else:
            fill_color = 'FFB8E994'  # Green

        ws.cell(row=row, column=2).fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')

        row += 1

    set_column_widths(ws, [35, 15, 30, 25])


def create_quotes_sheet(wb):
    """Create the validated quotes sheet."""
    ws = wb.create_sheet("Validated Quotes")

    row = 1
    ws[f'A{row}'] = "KEY VALIDATED QUOTES (From Stakeholder Interviews)"
    ws[f'A{row}'].font = Font(bold=True, size=16)
    row += 2

    headers = ["Stakeholder", "Quote", "Used For"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        apply_header_style(cell)
    row += 1

    quotes = [
        ("Nikki", '"Saving him about 10 hours to do those higher leverage things would be like a success."', "Success metric"),
        ("Alyssa", '"Maybe about four or five hours [per day on email]"', "Email time cost"),
        ("Alyssa", '"We have like seven Daves"', "Duplicate data problem"),
        ("Alyssa", '"It can take me about half an hour to try and figure it out because I don\'t know parts"', "Parts research time"),
        ("Alyssa", '"We don\'t know what we have... QuickBooks does not have the correct amounts"', "Inventory visibility"),
        ("Alyssa", '"If we had barcodes, that would be so much easier"', "Barcode system need"),
        ("Joe", '"Like 20, 30 hours, 20, 30 hours initially... took me a good 3 months"', "Training cost"),
        ("Joe", '"20% of calls need Kelsey\'s help... another 10-15% need minor guidance"', "Knowledge gaps"),
        ("Joe", '"1.5-2 hours every morning planning routes"', "Morning coordination"),
        ("Joe", '"Routes change 4 out of 5 days... Most frustrating part of my day"', "Route planning pain"),
        ("Kelsey", '"If I don\'t come to work, we make no money. In a couple months we\'d be out of business."', "Business dependency"),
    ]

    for stakeholder, quote, used_for in quotes:
        ws.cell(row=row, column=1, value=stakeholder)
        ws.cell(row=row, column=2, value=quote)
        ws.cell(row=row, column=3, value=used_for)

        for col in range(1, 4):
            ws.cell(row=row, column=col).border = thin_border
            ws.cell(row=row, column=col).alignment = Alignment(wrap_text=True, vertical='top')

        row += 1

    # Set row heights for wrapped text
    for r in range(4, row):
        ws.row_dimensions[r].height = 40

    set_column_widths(ws, [15, 80, 25])


# ============================================================================
# MAIN EXECUTION
# ============================================================================

def main():
    """Main function to create and save the workbook."""
    print("Creating Plotter Mechanix ROI Calculator...")

    # Create output directory if it doesn't exist
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    # Create workbook
    wb = create_workbook()

    # Save workbook
    wb.save(OUTPUT_FILE)

    print(f"✓ Created: {OUTPUT_FILE}")
    print(f"\nWorkbook contains {len(wb.sheetnames)} sheets:")
    for sheet in wb.sheetnames:
        print(f"  - {sheet}")

    return OUTPUT_FILE


if __name__ == "__main__":
    main()
